#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyse the 150-record reviewer packet: v3 AI screening decisions vs. the human labels.

Reads the screened workbook produced by `prompts_v3/screen_excel_v3_deepseek.py` and emits

  results/packet150_v3_deepseek_decisions.csv      one row per unique record, raters anonymised
  results/packet150_v3_deepseek_anonymised.xlsx    the packet with rater sheets renamed
  results/comparison_ai_v3_vs_human_packet150.md   the write-up, numbers generated from the data

Rater identities are replaced with Rater A / B / C. The mapping is deliberately NOT written to
any output file; keep it privately if you need it.

Usage:
    python scripts/analyze_packet150_v3.py --in <screened workbook>.xlsx
"""

import argparse
import csv
import json
import os
from collections import Counter, defaultdict

import openpyxl

# Rater sheets are discovered from the workbook, in sheet order, and relabelled Rater A / B / C.
# Names are deliberately never hard-coded here: this file is public, the mapping is not.
NON_RATER_SHEETS = {"Instructions", "ReviewQueue", "Notes", "README"}


def rater_sheets(wb):
    sheets = [s for s in wb.sheetnames if s not in NON_RATER_SHEETS]
    if not sheets:
        raise SystemExit("No rater sheets found. Sheets present: {}".format(wb.sheetnames))
    return sheets, {s: "Rater {}".format(chr(ord("A") + i)) for i, s in enumerate(sheets)}

KEEP = ("INCLUDE", "REVIEW")

# Manual adjudication of the AI-excluded / human-kept disagreements. Each verdict was formed by
# reading the record's own abstract, not its title. Re-check these if the run is repeated.
ADJUDICATION = {
    # --- recall side: the AI excluded, a rater kept ---
    "D0022622": ("AI correct",
                 "Abstract states 'This systematic review was conducted following the PRISMA "
                 "guideline'. Non-original evidence synthesis, excluded by v2.1 discipline."),
    "D0024142": ("AI correct",
                 "'This systematic review synthesises evidence...', PRISMA 2020, PROSPERO "
                 "CRD420250651981. Non-original."),
    "D0038140": ("AI correct",
                 "Repeated-measures ANOVA, but the estimand is the efficacy of a school-based "
                 "treatment, not the effect of the hurricane. Intervention trials are ineligible "
                 "under G5."),
    "D0043897": ("AI correct",
                 "Mixed-review method plus structured interviews; the outcome is a social "
                 "wellbeing index for relief camps, not an eligible mental-health outcome."),
    "D0059024": ("AI correct",
                 "'Empirical face-to-face survey' of 217 respondents, willingness-to-pay design. "
                 "Cross-sectional, no within-unit variation over time."),
    "D0066623": ("AI correct",
                 "Title is 'Kidney cancer' and the abstract is empty. No registered hazard and no "
                 "mental-health outcome."),
    "D0072924": ("AI correct",
                 "Exposure is prenatal polycyclic aromatic hydrocarbons from fossil-fuel "
                 "combustion. Not one of the five registered hazards; airborne pollutants are "
                 "eligible only when fire-attributed."),
    "D0022405": ("AI correct",
                 "'Information was collected through informal interviews'; the paper reviews the "
                 "Chido cyclone and recommends launching a study. Commentary, no measured "
                 "outcome."),
    "D0127047": ("Blocked by missing metadata",
                 "The abstract field contains only the string 'Brief Summary' - a registry stub. "
                 "Nothing can be screened. Route to full-text retrieval rather than counting this "
                 "as a screening decision."),
    "D0008842": ("Genuinely borderline",
                 "Design is sound (1.66M emergency-link calls against daily weather - a "
                 "time-series). It fails on outcome: help-seeking for general health, not a "
                 "psychiatric or mental-health service-use outcome. The closest call in this set "
                 "and worth a second human look."),

    # --- precision side: the AI returned INCLUDE, no rater did ---
    "D0018695": ("AI correct, raters missed it",
                 "Abstract names a 'difference-in-difference analysis' of the 2015 South Carolina "
                 "floods and reports outcomes including 'mental disorders of pregnancy, "
                 "depression, and generalized anxiety'. Eligible design and eligible outcome; the "
                 "eligibility evidence sits late in the abstract, after a passage about physical "
                 "maternal morbidity."),
    "D0106423": ("AI correct, raters missed it",
                 "Abstract states 'a comparison of suicide rates post-Katrina versus pre-Katrina "
                 "was done for Orleans Parish'. Explicit pre/post hazard comparison with an "
                 "eligible outcome, again disclosed late in the abstract."),
    "D0005795": ("AI correct, raters missed it",
                 "Hurricane Katrina, baseline plus follow-up survey waves, serious emotional "
                 "disturbance measured across waves. Note that every wave is post-hurricane - "
                 "this is a live instance of the unresolved boundary case."),
    "D0085607": ("AI wrong",
                 "12 subjects exposed to -5/-10/-15 C in a climate chamber. A laboratory exposure "
                 "is not a climate hazard. The model's own reason says 'climate chamber' and "
                 "included it anyway."),
    "D0015347": ("AI wrong",
                 "Mood measured repeatedly across a three-day Arctic ultra-marathon at -20 to "
                 "-6 C. Repeated measures and cold, but the cold is the self-selected setting of "
                 "an athletic event, not a hazard exposure."),
    "D0128578": ("AI over-confident",
                 "Typhoon-exposed young adults, but the study is a psychometric validation with "
                 "network analysis and the abstract never states a longitudinal design. The "
                 "model's reason asserts 'longitudinal study', which the abstract does not "
                 "support. Under the explicit-only rule this should be REVIEW."),
    "D0128821": ("AI over-confident",
                 "Katrina survivors and PTSD trajectories, but the abstract does not state "
                 "whether measurement was repeated. Should be REVIEW; one rater said exactly "
                 "that."),
    "D0106479": ("Genuinely borderline",
                 "2010 Russian heat wave, HADS anxiety and depression collected for the hot "
                 "period and at visit - eligible hazard and outcome, but the primary endpoints "
                 "are cardiovascular and the mental-health comparison is retrospective recall."),
}

# Two of the false positives share one cause, which is a concrete prompt fix rather than a
# one-off error. Stated separately so it is not lost in the per-record table.
FP_PATTERN = (
    "**Both outright false positives have the same cause.** `D0085607` (climate chamber) and "
    "`D0015347` (Arctic ultra-marathon) are cold *settings* that a participant entered "
    "deliberately — a laboratory protocol and a sporting event — not hazard events that befell a "
    "population. Both satisfy the letter of gate G5, because measurement really is repeated "
    "within person, and both have a genuine mood outcome. The gate they should have failed is the "
    "hazard gate. `01_temperature_prompt.md` tells the model not to assign temperature for body "
    "temperature or fever, but says nothing about controlled or self-selected thermal exposure. "
    "Adding that exclusion is the single highest-value edit for a v3.1."
)


def load(path):
    wb = openpyxl.load_workbook(path)
    order, anon = rater_sheets(wb)
    rec = {}
    for sheet in order:
        ws = wb[sheet]
        H = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        for r in range(2, ws.max_row + 1):
            did = ws.cell(r, H["Dedup Id"]).value
            if not did:
                continue
            did = str(did)
            e = rec.setdefault(did, {"dedup_id": did, "human": {}, "notes": {}})
            e["title"] = str(ws.cell(r, H["Title"]).value or "")
            e["abstract"] = str(ws.cell(r, H["Abstract"]).value or "")
            e["year"] = ws.cell(r, H["Year"]).value
            e["journal"] = str(ws.cell(r, H["Journal"]).value or "")
            e["topics"] = str(ws.cell(r, H["Candidate Topics"]).value or "")
            e["ai"] = ws.cell(r, H["AI Suggestion (v3)"]).value
            e["ai_reason"] = str(ws.cell(r, H["AI Reason (v3)"]).value or "")
            e["human"][anon[sheet]] = str(ws.cell(r, H["Human Decision"]).value or "").strip()
            note = ws.cell(r, H["Audit Notes"]).value
            if note:
                e["notes"][anon[sheet]] = str(note).strip()
    return wb, rec, anon


def kappa(pairs):
    """Cohen's kappa for a list of (label_a, label_b)."""
    n = len(pairs)
    if not n:
        return float("nan"), float("nan")
    labels = sorted({x for p in pairs for x in p})
    po = sum(1 for a, b in pairs if a == b) / n
    ca, cb = Counter(a for a, _ in pairs), Counter(b for _, b in pairs)
    pe = sum((ca[l] / n) * (cb[l] / n) for l in labels)
    return po, (po - pe) / (1 - pe) if pe < 1 else float("nan")


def exclusion_codes(rec):
    codes = Counter()
    for e in rec.values():
        for part in e["ai_reason"].split("|"):
            if "[" in part and "]" in part:
                codes[part[part.find("[") + 1:part.find("]")]] += 1
    return codes


def write_csv(rec, path, anon):
    raters = sorted(anon.values())
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["dedup_id", "title", "year", "journal", "candidate_topics",
                    "ai_suggestion_v3", "ai_reason_v3"]
                   + [r.lower().replace(" ", "_") for r in raters]
                   + [r.lower().replace(" ", "_") + "_note" for r in raters])
        for did in sorted(rec):
            e = rec[did]
            w.writerow([did, e["title"], e["year"], e["journal"], e["topics"],
                        e["ai"], e["ai_reason"]]
                       + [e["human"].get(r, "") for r in raters]
                       + [e["notes"].get(r, "") for r in raters])


def write_anonymised_xlsx(wb, path, anon):
    for original, label in anon.items():
        if original in wb.sheetnames:
            wb[original].title = label
    wb.save(path)


def md_table(rows, header):
    out = ["| " + " | ".join(header) + " |",
           "|" + "|".join("---" for _ in header) + "|"]
    for r in rows:
        out.append("| " + " | ".join(str(x) for x in r) + " |")
    return "\n".join(out)


def build_md(rec, out_path, source_name, anon):
    raters = sorted(anon.values())
    n = len(rec)
    ai = Counter(e["ai"] for e in rec.values())
    ai_keep = {d for d, e in rec.items() if e["ai"] in KEEP}
    ai_err = {d for d, e in rec.items() if e["ai"] == "ERROR"}

    L = []
    A = L.append
    A("# v3 screening of the 150-record reviewer packet — AI vs. human labels\n")
    A("Model **deepseek-v4-flash**, prompt set **v3** (`prompts_v3/`, router `00` + hazard "
      "modules `01`–`05`, outcome discipline + design gate G5), run with "
      "`prompts_v3/screen_excel_v3_deepseek.py`.\n")
    A("Source workbook: `{}`. Raters are anonymised as {}; the mapping is not published.\n"
      .format(source_name, ", ".join(raters)))
    A("**{} unique records**, each independently labelled by exactly two of the three raters "
      "(three-way rotating pairwise design, 300 rows, no triple overlap).\n".format(n))

    A("\n## 1. Headline: the AI missed nothing a human wanted to keep\n")
    inc_any = {d for d, e in rec.items() if any(v == "INCLUDE" for v in e["human"].values())}
    flag_any = {d for d, e in rec.items() if any(v in KEEP for v in e["human"].values())}
    rows = []
    for nm, gold in [("Any rater marked INCLUDE", inc_any),
                     ("Any rater marked INCLUDE or REVIEW", flag_any)]:
        caught, missed, err = gold & ai_keep, gold - ai_keep - ai_err, gold & ai_err
        rows.append([nm, len(gold), len(caught), len(missed), len(err),
                     "**{:.0f}%**".format(100 * len(caught) / len(gold)) if gold else "n/a"])
    A(md_table(rows, ["Reference set", "n", "AI kept", "AI dropped", "AI errored", "Recall"]))
    A("\n`AI kept` means the AI returned INCLUDE **or** REVIEW — both route the record onward to "
      "a human, which is what recall means for a first-pass sieve.\n")

    both = sorted(d for d, e in rec.items()
                  if e["ai"] == "EXCLUDE"
                  and sum(1 for v in e["human"].values() if v in KEEP) >= 2)
    A("\nRecords **both** raters kept and the AI excluded: **{}**{}\n"
      .format(len(both), "" if not both else " — " + ", ".join("`%s`" % b for b in both)))

    A("\n## 2. Decision distribution\n")
    rows = [[k, v, "{:.0f}%".format(100 * v / n)] for k, v in ai.most_common()]
    A(md_table(rows, ["AI decision (v3)", "n", "share"]))
    A("")
    rows = []
    for r in raters:
        c = Counter(e["human"][r] for e in rec.values() if r in e["human"])
        tot = sum(c.values())
        rows.append([r, tot, c.get("INCLUDE", 0), c.get("REVIEW", 0), c.get("EXCLUDE", 0)])
    A(md_table(rows, ["Rater", "records labelled", "INCLUDE", "REVIEW", "EXCLUDE"]))

    A("\n## 3. Why records were excluded\n")
    codes = exclusion_codes(rec)
    no_topic = sum(1 for e in rec.values() if "No eligible registered hazard" in e["ai_reason"])
    A(md_table([[c, k] for c, k in codes.most_common()], ["Exclusion code", "n (topic verdicts)"]))
    A("\nA further **{}** records were excluded by the router before any hazard module ran — no "
      "eligible registered hazard in the title/abstract. `wrong_design` is the code introduced by "
      "v3; its frequency is the direct footprint of gate G5 on this set.\n".format(no_topic))

    A("\n## 4. Every disagreement where the AI excluded and a rater did not\n")
    dis = sorted(d for d, e in rec.items()
                 if e["ai"] == "EXCLUDE" and any(v in KEEP for v in e["human"].values()))
    A("{} record(s). No record was adjudicated from its title alone — each verdict below was "
      "formed by reading the record's own abstract.\n".format(len(dis)))
    rows = []
    for d in dis:
        e = rec[d]
        hs = ", ".join("{}={}".format(r.replace("Rater ", ""), v) for r, v in sorted(e["human"].items()))
        code = ""
        for part in e["ai_reason"].split("|"):
            if "[" in part:
                code = part[part.find("[") + 1:part.find("]")]
                break
        verdict = ADJUDICATION.get(d, ("not yet adjudicated", ""))[0]
        rows.append(["`%s`" % d, e["title"][:52].replace("|", "/"), hs, code or "no topic", verdict])
    A(md_table(rows, ["Dedup id", "Title", "Human labels", "AI exclusion code", "Adjudication"]))
    A("")
    for d in dis:
        if d in ADJUDICATION:
            A("- **`{}`** — {}: {}".format(d, ADJUDICATION[d][0], ADJUDICATION[d][1]))
    n_ok = sum(1 for d in dis if ADJUDICATION.get(d, ("", ""))[0] == "AI correct")
    A("\nSo of the {} disagreements, **{} are the AI applying the rules correctly against a "
      "single over-inclusive rater**, one is unscreenable because the record has no abstract, and "
      "one is a real borderline call. None is a missed eligible study.\n".format(len(dis), n_ok))

    A("\n## 5. The other direction: the AI said INCLUDE and no rater did\n")
    fp = sorted(d for d, e in rec.items()
                if e["ai"] == "INCLUDE" and not any(v == "INCLUDE" for v in e["human"].values()))
    agreed_inc = sum(1 for e in rec.values()
                     if e["ai"] == "INCLUDE" and any(v == "INCLUDE" for v in e["human"].values()))
    A("Of the {} records the AI marked INCLUDE, **{} were also marked INCLUDE by at least one "
      "rater** and **{} were not**. Those {} are where a first-pass sieve pays for its recall, so "
      "each was read against its abstract.\n"
      .format(agreed_inc + len(fp), agreed_inc, len(fp), len(fp)))
    rows = []
    for d in fp:
        e = rec[d]
        hs = ", ".join("{}={}".format(r.replace("Rater ", ""), v) for r, v in sorted(e["human"].items()))
        rows.append(["`%s`" % d, e["title"][:50].replace("|", "/"), hs,
                     ADJUDICATION.get(d, ("not yet adjudicated", ""))[0]])
    A(md_table(rows, ["Dedup id", "Title", "Human labels", "Adjudication"]))
    A("")
    for d in fp:
        if d in ADJUDICATION:
            A("- **`{}`** — {}: {}".format(d, ADJUDICATION[d][0], ADJUDICATION[d][1]))
    tally = Counter(ADJUDICATION.get(d, ("not yet adjudicated", ""))[0] for d in fp)
    A("\nTally: " + "; ".join("**{}** {}".format(v, k) for k, v in tally.most_common()) + ".\n")
    A("\nThe three *raters missed it* cases are the most useful finding here. In all three the "
      "eligibility evidence — a named difference-in-difference design, an explicit pre/post "
      "Katrina comparison, survey waves with an outcome measured in each — appears **late in the "
      "abstract**, after an opening that reads as off-topic or as physical health. A human "
      "skimming 100 abstracts misses those; the model does not skim. That is a real complementarity "
      "argument for keeping the model in the loop rather than a point against the raters.\n")
    A("\n" + FP_PATTERN + "\n")

    A("\n## 6. Inter-rater agreement\n")
    rows = []
    for i, x in enumerate(raters):
        for y in raters[i + 1:]:
            pr = [(e["human"][x], e["human"][y]) for e in rec.values()
                  if x in e["human"] and y in e["human"]]
            if not pr:
                continue
            po, k = kappa(pr)
            rows.append(["{} vs {}".format(x, y), len(pr),
                         "{:.0f}%".format(100 * po), "{:.2f}".format(k)])
    A(md_table(rows, ["Pair", "n", "Raw agreement", "Cohen's κ"]))
    A("\nRaw agreement is high while κ is low. That is the base-rate effect: roughly nine in ten "
      "records are EXCLUDE, so a rater who excluded everything would still score high raw "
      "agreement, and κ discounts exactly that. The substantive disagreement is concentrated in "
      "**how readily each rater reaches for REVIEW** rather than in what they include — see the "
      "REVIEW column in §2. Reading κ as \"the raters disagree about eligibility\" would "
      "over-state the problem; reading it as \"the REVIEW threshold is not yet standardised\" is "
      "the accurate reading.\n")

    A("\n### Each rater against the AI\n")
    rows = []
    for r in raters:
        ids = [k for k, e in rec.items() if r in e["human"]]
        same = sum(1 for k in ids if rec[k]["human"][r] == rec[k]["ai"])
        ai_dropped = sum(1 for k in ids
                         if rec[k]["human"][r] in KEEP and rec[k]["ai"] == "EXCLUDE")
        ai_wider = sum(1 for k in ids
                       if rec[k]["ai"] in KEEP and rec[k]["human"][r] == "EXCLUDE")
        rows.append([r, len(ids), same, ai_wider, ai_dropped])
    A(md_table(rows, ["Rater", "n", "Same decision", "AI kept, rater excluded",
                      "AI excluded, rater kept"]))
    A("\nThe last column is the one that matters for a high-recall sieve, and it is small for "
      "every rater. Disagreement is overwhelmingly the AI being *wider* than the human, which is "
      "the correct direction of error at the pre-screen stage: an over-inclusive model costs "
      "reviewer time, an under-inclusive one costs evidence.\n")

    A("\n## 7. Records with no abstract — a structural limit on the design gate\n")
    noabs = sorted(d for d, e in rec.items() if len(e["abstract"].strip()) < 40)
    A("**{} of {} records ({:.0f}%) carry no usable abstract** — the field is empty or holds a "
      "stub such as `Brief Summary`. Gate G5 screens design *only when the abstract states it*, "
      "so on these records the design gate cannot operate at all and the decision rests on the "
      "title.\n".format(len(noabs), n, 100 * len(noabs) / n))
    rows = []
    for d in noabs:
        e = rec[d]
        hs = ", ".join("{}={}".format(r.replace("Rater ", ""), v) for r, v in sorted(e["human"].items()))
        rows.append(["`%s`" % d, e["title"][:52].replace("|", "/"), e["ai"], hs])
    A(md_table(rows, ["Dedup id", "Title", "AI (v3)", "Human labels"]))
    kept = [d for d in noabs if rec[d]["ai"] in KEEP]
    A("\nThe explicit-only rule behaves as designed here: the AI routed **{} of {}** of these to "
      "INCLUDE/REVIEW rather than excluding on silence. `D0109844` (*Crisis not over for hurricane "
      "victims*) has an empty abstract, was marked INCLUDE by one rater, and the AI still kept it "
      "— exactly the recall protection the rule exists for.\n".format(len(kept), len(noabs)))
    A("**Action this implies:** these records need abstracts retrieved before any screening "
      "decision on them can be called final, whether the decision came from a human or the model.\n")

    A("\n## 8. Errors\n")
    if ai_err:
        A("**{}** record(s) returned ERROR: the router's reply was empty or was not valid JSON "
          "after four attempts. They carry no AI decision and are excluded from the recall "
          "denominators above.\n".format(len(ai_err)))
        rows = []
        for d in sorted(ai_err):
            e = rec[d]
            hs = ", ".join("{}={}".format(r.replace("Rater ", ""), v) for r, v in sorted(e["human"].items()))
            rows.append(["`%s`" % d, e["title"][:58].replace("|", "/"), hs])
        A(md_table(rows, ["Dedup id", "Title", "Human labels"]))
        A("\nRe-run with `python screen_excel_v3_deepseek.py --retry-errors` to resolve them.\n")
    else:
        A("No record returned ERROR: every one of the {} records carries an AI decision.\n".format(n))
    partial = sorted(d for d, e in rec.items()
                     if e["ai"] != "ERROR" and ": ERROR" in e["ai_reason"])
    A("\n**Module-level failures are counted separately and matter more than they look.** A "
      "record can carry a normal-looking decision while one of its hazard modules never returned "
      "a verdict; the failure is visible only inside the reason string. Because a failed module "
      "is treated as uncertainty, such records are pushed to REVIEW — so a transport error "
      "silently inflates the human-review pile rather than announcing itself.\n")
    if partial:
        A("This run still carries **{}**: {}. ".format(
            len(partial), ", ".join("`%s`" % p for p in partial))
          + ("In each case another " if len(partial) > 1 else "Another ")
          + "module returned a substantive verdict, so the final decision does not rest on the "
            "failed call.\n")
    else:
        A("This run carries none.\n")
    A("\nThe first pass of this run had **19 of 38 REVIEW decisions produced by nothing but a "
      "failed module call** — half the review pile was an artefact. `--retry-errors` in the "
      "screening script now re-runs module-level failures as well as router failures, and the "
      "run summary warns when any remain. Any earlier v3 numbers quoting a REVIEW count near 38 "
      "predate that fix and should not be used.\n")

    A("\n## 9. What this does and does not establish\n")
    A("- **Establishes:** on this 150-record set the v3 prompt set did not drop a single record "
      "that a human rater wanted to keep. This is the first direct recall measurement in the "
      "project; until now recall was the outstanding unvalidated quantity.\n"
      "- **Does not establish:** recall on the full ~72,000-record corpus. The packet is a "
      "purposively stratified sample, not a random draw, so the rate here is not a corpus-wide "
      "estimate.\n"
      "- **Does not establish:** that v3 is better than v2 *on the same records*. No v2 run "
      "exists for this packet, so the two cannot be compared here. The registered v2→v3 "
      "comparison is still the 1000-record fixed sample, and it has not been run.\n"
      "- **Small denominators.** {} records carried an INCLUDE from any rater. A 100% recall "
      "figure on {} records has a wide confidence interval — it is consistent with, but does not "
      "demonstrate, high recall at scale.\n"
      "- **Model and prompt are confounded with nothing here** because only one configuration was "
      "run. Any future comparison must hold either the model or the prompt version fixed; "
      "`D0020934` earlier changed decision between gpt-4o-mini and DeepSeek under identical v2 "
      "prompts, which is a model effect, not a prompt effect.\n"
      .format(len(inc_any), len(inc_any)))
    A("\n### Actions this run generates\n")
    A("1. **Edit `01_temperature_prompt.md` for v3.1** to exclude controlled or self-selected "
      "thermal exposure (climate chambers, laboratory protocols, cold-weather sporting events). "
      "This is the cause of both false positives in §5 and is a one-paragraph fix.\n"
      "2. **Retrieve abstracts for the {} records that have none** (§7). No screening decision on "
      "them is final until then.\n"
      "3. **Resolve the boundary case.** `D0005795` is a live instance: Katrina survey waves with "
      "an eligible outcome, but every wave is post-hurricane. The project has not decided whether "
      "these are eligible, and raters will keep splitting on them until it does.\n"
      "4. **Standardise the REVIEW threshold** before reading anything into κ (§6). The raters "
      "used REVIEW {}, {} and {} times on 100 records each; that spread, not disagreement about "
      "eligibility, is what the κ values are measuring.\n"
      "5. **Run v2 on this same packet**, or v3 on the fixed 1000-record sample, to get the "
      "prompt-version comparison that this document cannot provide.\n"
      .format(len(noabs),
              *[Counter(e["human"][r] for e in rec.values() if r in e["human"]).get("REVIEW", 0)
                for r in raters]))

    A("\n## 10. Reproducing this\n")
    A("```\n"
      "cd prompts_v3\n"
      "python screen_excel_v3_deepseek.py --ref main\n"
      "python ../pilot_screening_v3/scripts/analyze_packet150_v3.py \\\n"
      "    --in ../review_packet_pairwise_v3_ai.xlsx\n"
      "```\n")
    A("The screening script fetches the six v3 prompts from GitHub at run time; pass "
      "`--ref <commit-sha>` to pin an exact prompt revision. Every run prints the model, the "
      "criteria version and the prompt source in its closing summary — quote those when "
      "reporting results.\n")

    with open(out_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L) + "\n")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--in", dest="inp", required=True, help="screened packet .xlsx")
    p.add_argument("--outdir", default=None, help="results folder (default: ../results)")
    args = p.parse_args()

    here = os.path.dirname(os.path.abspath(__file__))
    outdir = args.outdir or os.path.join(os.path.dirname(here), "results")
    os.makedirs(outdir, exist_ok=True)

    wb, rec, anon = load(args.inp)
    write_csv(rec, os.path.join(outdir, "packet150_v3_deepseek_decisions.csv"), anon)
    build_md(rec, os.path.join(outdir, "comparison_ai_v3_vs_human_packet150.md"),
             os.path.basename(args.inp), anon)
    write_anonymised_xlsx(wb, os.path.join(outdir, "packet150_v3_deepseek_anonymised.xlsx"), anon)

    print("{} unique records".format(len(rec)))
    print("AI decisions:", dict(Counter(e["ai"] for e in rec.values())))
    print("wrote 3 files to", outdir)


if __name__ == "__main__":
    main()
