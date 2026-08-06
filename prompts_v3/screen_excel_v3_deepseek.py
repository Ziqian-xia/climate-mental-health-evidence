#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Screen the reviewer packet (Ziqian / Jacob / Tony sheets) with the v3 prompts on DeepSeek.

The six v3 prompts are fetched from GitHub by default, so a collaborator can clone this
repository (or just download this single file) and run it without editing any paths. If the
network is unavailable the script falls back to a local `prompts_v3/` folder and says so.

Title and abstract are read directly from the Excel file, so no corpus download is needed.
Two new columns per sheet - "AI Suggestion (v3)" and "AI Reason (v3)" - are written into a
NEW Excel file. The existing "Human Decision" and "Audit Notes" columns are never modified:
the packet exists to produce independent human labels, and contaminating those columns with
model output would make the validation circular.

Records that appear on more than one sheet (the pairwise assignment) are screened once and
the same result is written to every row carrying that dedup id.

Progress is checkpointed to JSON after every record, so an interrupted run resumes on restart.

Usage
-----
    python screen_excel_v3_deepseek.py                     # normal run, prompts from GitHub
    python screen_excel_v3_deepseek.py --limit 3           # 3-record smoke test first
    python screen_excel_v3_deepseek.py --prompts-dir ./prompts_v3    # force local prompts
    python screen_excel_v3_deepseek.py --ref <commit-sha>  # pin an exact prompt revision
    python screen_excel_v3_deepseek.py --in packet.xlsx --out packet_v3.xlsx

Requirements
------------
    pip install openpyxl openai
    set DEEPSEEK_API_KEY=...        (or paste the key when prompted)
"""

import argparse
import getpass
import json
import os
import re
import sys
import time
from collections import Counter

# ------------------------- configuration -------------------------
REPO = "Ziqian-xia/climate-mental-health-evidence"
DEFAULT_REF = "main"
PROMPT_DIR_IN_REPO = "prompts_v3"          # root-level in this repo; no subfolder

# logical name -> file name inside PROMPT_DIR_IN_REPO
PROMPT_FILES = {
    "router":      "00_candidate_topics_prompt.md",
    "temperature": "01_temperature_prompt.md",
    "wildfire":    "02_wildfire_prompt.md",
    "flood":       "03_flood_prompt.md",
    "cyclone":     "04_cyclone_prompt.md",
    "drought":     "05_drought_prompt.md",
}

# Every v3 prompt, including the router, carries this marker. Used as a version guard so a
# stale or wrong-version file can never be screened with silently.
VERSION_MARKER = "**Version: v3**"

DEFAULT_EXCEL_NAME = "review_packet_pairwise.xlsx"
DEFAULT_SHEETS = ["Ziqian", "Jacob", "Tony"]

MODEL = "deepseek-v4-flash"
BASE_URL = "https://api.deepseek.com"
MAX_TOKENS = 800
PAUSE_SEC = 0.2
HTTP_TIMEOUT = 30
CRITERIA_VERSION = "per-topic prompt set 00-05 v3 (outcome + design discipline)"
# -----------------------------------------------------------------

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def log(m):
    print(m, flush=True)


def die(m):
    print("\n[ERROR] " + m + "\n", flush=True)
    sys.exit(1)


try:
    import openpyxl
except ImportError:
    die("Missing openpyxl. Run: pip install openpyxl openai")
try:
    from openai import OpenAI
except ImportError:
    die("Missing openai. Run: pip install openpyxl openai")


# ------------------------- prompt loading -------------------------

def raw_url(ref, filename):
    return "https://raw.githubusercontent.com/{}/{}/{}/{}".format(
        REPO, ref, PROMPT_DIR_IN_REPO, filename)


def fetch_prompts_from_github(ref):
    """Download the six v3 prompts. Raises on any failure."""
    import urllib.request
    prompts = {}
    for name, filename in PROMPT_FILES.items():
        url = raw_url(ref, filename)
        req = urllib.request.Request(url, headers={"User-Agent": "v3-screening-script"})
        with urllib.request.urlopen(req, timeout=HTTP_TIMEOUT) as r:
            if getattr(r, "status", 200) != 200:
                raise RuntimeError("HTTP {} for {}".format(r.status, url))
            prompts[name] = r.read().decode("utf-8")
    return prompts


def find_local_prompt_dir(explicit):
    """Return a directory that contains all six prompt files, or None."""
    candidates = []
    if explicit:
        candidates.append(explicit)
    else:
        candidates += [
            os.path.join(SCRIPT_DIR, PROMPT_DIR_IN_REPO),
            SCRIPT_DIR,                                        # script sitting inside prompts_v3/
            os.path.join(os.path.dirname(SCRIPT_DIR), PROMPT_DIR_IN_REPO),
            os.path.join(os.getcwd(), PROMPT_DIR_IN_REPO),
            os.getcwd(),
        ]
    for d in candidates:
        if d and all(os.path.exists(os.path.join(d, f)) for f in PROMPT_FILES.values()):
            return os.path.abspath(d)
    return None


def load_prompts_from_dir(d):
    prompts = {}
    for name, filename in PROMPT_FILES.items():
        with open(os.path.join(d, filename), encoding="utf-8") as fh:
            prompts[name] = fh.read()
    return prompts


def check_version(prompts, source):
    bad = [n for n, text in prompts.items() if VERSION_MARKER not in text]
    if bad:
        die("These prompt files do not look like v3 (missing the '{}' marker): {}\n"
            "Source was: {}\nRefusing to run - the output would be mislabelled as v3."
            .format(VERSION_MARKER, ", ".join(sorted(bad)), source))


def load_prompts(args):
    """GitHub first, local folder as a fallback. Returns (prompts, source_description)."""
    local_dir = find_local_prompt_dir(args.prompts_dir)

    if args.prompts_dir:
        if not local_dir:
            die("--prompts-dir was given but that folder does not contain all six prompt "
                "files (00-05):\n  {}".format(args.prompts_dir))
        prompts = load_prompts_from_dir(local_dir)
        source = "local folder {} (forced with --prompts-dir)".format(local_dir)
    else:
        try:
            prompts = fetch_prompts_from_github(args.ref)
            source = "GitHub {}@{}/{}".format(REPO, args.ref, PROMPT_DIR_IN_REPO)
        except Exception as e:
            log("      Could not reach GitHub: {}".format(str(e)[:160]))
            if not local_dir:
                die("GitHub is unreachable and no local copy of the six v3 prompts was found.\n"
                    "Either connect to the network (a phone hotspot usually works), or pass\n"
                    "  --prompts-dir <folder containing 00_..05_ prompt .md files>")
            log("      Falling back to the local copy.")
            prompts = load_prompts_from_dir(local_dir)
            source = "local folder {} (GitHub unreachable)".format(local_dir)

    check_version(prompts, source)
    log("      Prompt source: {}".format(source))
    log("      Loaded {} prompts; temperature prompt = {} chars "
        "(v3 ~8829, v2 ~5900, v1 ~2600).".format(len(prompts), len(prompts["temperature"])))
    return prompts, source


# ------------------------- model calls -------------------------

def extract_json(text):
    if not text:
        raise ValueError("empty reply")
    t = text.strip()
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", t, re.S)
    if m:
        t = m.group(1)
    else:
        a, b = t.find("{"), t.rfind("}")
        if a != -1 and b != -1 and b > a:
            t = t[a:b + 1]
    return json.loads(t)


def get_api_key():
    key = os.environ.get("DEEPSEEK_API_KEY", "").strip()
    if key:
        return key
    log("\nNo DEEPSEEK_API_KEY detected.")
    log("Paste your DeepSeek API key (from platform.deepseek.com; it will NOT be shown as you type):")
    key = getpass.getpass("API key: ").strip()
    if not key:
        die("No API key entered.")
    return key


def verify_api_key(client, model):
    log("      Verifying API key ...")
    try:
        client.chat.completions.create(model=model, max_tokens=1,
                                       messages=[{"role": "user", "content": "ping"}])
        log("      API key is valid.")
    except Exception as e:
        msg = str(e)
        if any(t in msg for t in ("401", "invalid_api_key", "Unauthorized")):
            die("API key is invalid or unauthorized.\nOriginal error: " + msg[:200])
        if any(t in msg for t in ("429", "insufficient", "quota")):
            die("The API key works but has insufficient credit.\nOriginal error: " + msg[:200])
        die("API call failed (network or model-name issue): " + msg[:200])


def call_model(client, model, system_prompt, record):
    user_text = ("Here is one record to screen. Use ONLY its title and abstract.\n\n"
                 + json.dumps(record, ensure_ascii=False)
                 + "\n\nReturn exactly one JSON object as specified. Output JSON only.")
    last = None
    for attempt in range(4):
        try:
            resp = client.chat.completions.create(
                model=model, max_tokens=MAX_TOKENS, temperature=0,
                response_format={"type": "json_object"},
                messages=[{"role": "system", "content": system_prompt},
                          {"role": "user", "content": user_text}])
            return extract_json(resp.choices[0].message.content)
        except Exception as e:
            last = e
            time.sleep(2 * (attempt + 1))
    return {"_error": str(last)}


def screen_one(client, model, prompts, record):
    """Router assigns hazard topics, then one prompt per topic decides. Returns (decision, reason)."""
    router = call_model(client, model, prompts["router"], record)
    if "_error" in router:
        return "ERROR", "router failed: " + str(router["_error"])[:150]

    topics = router.get("candidate_topics", []) or []
    if not topics:
        return "EXCLUDE", "No eligible registered hazard identified."

    per, any_inc, inc_rev, any_rev = [], False, False, False
    for tp in topics:
        if tp not in prompts:
            continue
        out = call_model(client, model, prompts[tp], record)
        if "_error" in out:
            per.append("{}: ERROR".format(tp))
            any_rev = True
            continue
        dec = str(out.get("decision", "")).upper()
        rf = bool(out.get("review_flag", False))
        code = out.get("exclusion_code", "")
        why = str(out.get("one_line_reason", ""))[:150]
        per.append("{}: {}{} - {}".format(
            tp, dec, " [{}]".format(code) if code and code != "NA" else "", why))
        if dec == "INCLUDE":
            any_inc = True
            if rf:
                inc_rev = True
        elif dec == "MAYBE":
            any_rev = True
        if rf:
            any_rev = True

    reason = " | ".join(per)
    if any_inc:
        return ("REVIEW" if inc_rev else "INCLUDE"), reason
    if any_rev:
        return "REVIEW", reason
    return "EXCLUDE", reason


# ------------------------- excel helpers -------------------------

def find_default_excel():
    for d in (SCRIPT_DIR, os.path.dirname(SCRIPT_DIR), os.getcwd()):
        p = os.path.join(d, DEFAULT_EXCEL_NAME)
        if os.path.exists(p):
            return p
    return None


def header_map(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def ensure_column(ws, headers, title):
    """Return the column index for `title`, appending a new column if it is absent."""
    if title in headers and headers[title]:
        return headers[title]
    col = ws.max_column + 1
    ws.cell(1, col, title)
    headers[title] = col
    return col


# ------------------------- main -------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Screen the reviewer packet with v3 prompts on DeepSeek.")
    p.add_argument("--in", dest="inp", default=None,
                   help="input .xlsx (default: {} next to this script or its parent folder)"
                        .format(DEFAULT_EXCEL_NAME))
    p.add_argument("--out", dest="out", default=None,
                   help="output .xlsx (default: <input>_v3_ai.xlsx)")
    p.add_argument("--prompts-dir", default=None,
                   help="use local prompt files instead of GitHub")
    p.add_argument("--ref", default=DEFAULT_REF,
                   help="git ref or commit sha to fetch prompts from (default: %(default)s)")
    p.add_argument("--sheets", default=",".join(DEFAULT_SHEETS),
                   help="comma-separated sheet names (default: %(default)s)")
    p.add_argument("--model", default=MODEL, help="model name (default: %(default)s)")
    p.add_argument("--limit", type=int, default=0,
                   help="screen at most N unique records - use a small number for a smoke test")
    p.add_argument("--checkpoint", default=None,
                   help="resume file (default: screen_v3_progress.json next to the output)")
    return p.parse_args()


def main():
    args = parse_args()
    sheets = [s.strip() for s in args.sheets.split(",") if s.strip()]

    log("=" * 72)
    log("  Reviewer packet screening - v3 prompts on DeepSeek")
    log("=" * 72)

    excel_in = args.inp or find_default_excel()
    if not excel_in:
        die("Could not find {}.\nLooked next to this script, in its parent folder, and in the\n"
            "current directory. Pass the path explicitly:  --in <path to .xlsx>"
            .format(DEFAULT_EXCEL_NAME))
    excel_in = os.path.abspath(excel_in)
    if not os.path.exists(excel_in):
        die("Excel file not found:\n  " + excel_in)

    excel_out = os.path.abspath(args.out) if args.out else \
        os.path.splitext(excel_in)[0] + "_v3_ai.xlsx"
    checkpoint = os.path.abspath(args.checkpoint) if args.checkpoint else \
        os.path.join(os.path.dirname(excel_out), "screen_v3_progress.json")

    log("\n[1/5] Reading the Excel file ...")
    log("      in  : " + excel_in)
    log("      out : " + excel_out)
    wb = openpyxl.load_workbook(excel_in)
    missing = [s for s in sheets if s not in wb.sheetnames]
    if missing:
        die("Sheet(s) not found: {}\nSheets present: {}".format(missing, wb.sheetnames))

    records, locations = {}, {}
    for sheet in sheets:
        ws = wb[sheet]
        H = header_map(ws)
        for col in ("Dedup Id", "Title", "Abstract"):
            if col not in H:
                die("Sheet '{}' has no column named '{}'. Columns: {}".format(sheet, col, list(H)))
        for r in range(2, ws.max_row + 1):
            did = ws.cell(r, H["Dedup Id"]).value
            if not did:
                continue
            did = str(did)
            title = ws.cell(r, H["Title"]).value or ""
            abstract = ws.cell(r, H["Abstract"]).value or ""
            records.setdefault(did, {"dedup_id": did,
                                     "title": str(title),
                                     "abstract": str(abstract)})
            locations.setdefault(did, []).append((sheet, r))
    total_rows = sum(len(v) for v in locations.values())
    log("      {} rows across {} sheet(s); {} unique records."
        .format(total_rows, len(sheets), len(records)))

    log("\n[2/5] Loading v3 prompts ...")
    prompts, prompt_source = load_prompts(args)

    log("\n[3/5] Connecting to DeepSeek ...")
    client = OpenAI(api_key=get_api_key(), base_url=BASE_URL)
    verify_api_key(client, args.model)

    done = {}
    if os.path.exists(checkpoint):
        try:
            with open(checkpoint, encoding="utf-8") as fh:
                done = json.load(fh)
            log("      Resuming: {} record(s) already screened (from {}).".format(len(done), checkpoint))
        except Exception:
            done = {}

    todo = [d for d in records if d not in done]
    if args.limit:
        todo = todo[:args.limit]
        log("      --limit {} -> screening only {} record(s) this run.".format(args.limit, len(todo)))

    log("\n[4/5] Screening {} record(s) ...".format(len(todo)))
    t0 = time.time()
    for i, did in enumerate(todo, 1):
        dec, reason = screen_one(client, args.model, prompts, records[did])
        done[did] = {"decision": dec, "reason": reason}
        with open(checkpoint, "w", encoding="utf-8") as fh:
            json.dump(done, fh, ensure_ascii=False)
        if i % 5 == 0 or i == len(todo):
            el = time.time() - t0
            rate = el / i
            log("      {}/{} done  ({:.1f} min elapsed, ~{:.1f} min left)"
                .format(i, len(todo), el / 60, rate * (len(todo) - i) / 60))
        time.sleep(PAUSE_SEC)

    log("\n[5/5] Writing results into a new Excel file ...")
    for sheet in sheets:
        ws = wb[sheet]
        H = header_map(ws)
        c_dec = ensure_column(ws, H, "AI Suggestion (v3)")
        c_rea = ensure_column(ws, H, "AI Reason (v3)")
        for r in range(2, ws.max_row + 1):
            did = ws.cell(r, H["Dedup Id"]).value
            if did and str(did) in done:
                ws.cell(r, c_dec, done[str(did)]["decision"])
                ws.cell(r, c_rea, done[str(did)]["reason"][:900])
    wb.save(excel_out)

    cnt = Counter(v["decision"] for v in done.values())
    log("\n" + "=" * 72)
    log("  SUMMARY")
    log("=" * 72)
    log("  Unique records screened : {}".format(len(done)))
    for k, v in cnt.most_common():
        log("    {:<10} {}".format(k, v))
    log("  Model                   : {}".format(args.model))
    log("  Criteria                : {}".format(CRITERIA_VERSION))
    log("  Prompt source           : {}".format(prompt_source))
    log("  Runtime                 : {:.1f} min".format((time.time() - t0) / 60))
    log("\n  Saved: " + excel_out)
    log("  ('Human Decision' and 'Audit Notes' columns were not modified.)")
    log("  Delete {} once you are happy with the output.".format(checkpoint))
    if len(done) < len(records):
        log("\n  NOTE: {} of {} unique records are still unscreened (--limit was used)."
            .format(len(records) - len(done), len(records)))
        log("  Run the script again without --limit to finish; it resumes from the checkpoint.")


if __name__ == "__main__":
    main()
